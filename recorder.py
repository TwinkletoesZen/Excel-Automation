from openpyxl import Workbook, load_workbook
from datetime import date


wb = Workbook()
wb = load_workbook("Olympic.xlsx")

ws_basketball = wb["Basketball"]
ws_volleyball = wb["Volleyball"]
ws_soccer = wb["Soccer"]
ws_ping_pong = wb["Ping-Pong"]
ws_badminton = wb["Badminton"]
ws_pool = wb["Pool"]
# ws_standing = wb["Standing"]
# ws_All_House_NumWin = wb["All House"]
# ws_All_House_NumPoint = wb["Game Points"]

Unity = "UNITY"
Courage = "COURAGE"
Discovery = "DISCOVERY"
Legacy = "LEGACY"
Spirit = "SPIRIT"
Harmony = "HARMONY"

What_Game = input("What Game is it?: ").upper()



def Record_Basketball_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ").upper()
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  # if Winner == Unity:
  #   with open("House_Points/UnityP.txt", a):

  #   with open("House_Wins/Unity.txt", a):

  # if Loser == Unity:
  #   with open("House_Points/UnityP.txt", a):

  #   with open("House_Wins/Unity.txt", a):


  

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_basketball.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])    

def Record_Volleyball_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ")
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_volleyball.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])   
 
def Record_Soccer_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ")
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_soccer.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])    

def Record_PingPong_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ")
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_ping_pong.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])    

def Record_Badminton_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ")
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_badminton.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])    

def Record_Pool_Game():
  Between_Who = input("What House is it Between?: ")
  Winner = input("Winner: ")
  Loser = input("Loser: ")
  Points_Winner = input("Points for Winner: ")
  Points_Loser = input("Points for Loser: ")
  ask_Date = input("Is the game happening today?: ").upper()

  
  if ask_Date == "Y":
    Date = date.today()
  elif ask_Date == "N":
    Date = input("Enter the Date (yyyy/mm/dd)")

    
  ws_pool.append([Between_Who, Winner, Loser, Points_Winner, Points_Loser, Date])    



if What_Game == "BASKETBALL":
  Record_Basketball_Game()
elif What_Game == "VOLLEYBALL":
  Record_Volleyball_Game()
elif What_Game == "SOCCER":
  Record_Soccer_Game()
elif What_Game == "PINGPONG":
  Record_PingPong_Game()
elif What_Game == "BADMINTON":
  Record_Badminton_Game()
elif What_Game == "POOL":
  Record_Pool_Game()



wb.save("Olympic.xlsx")